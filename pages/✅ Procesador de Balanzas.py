

import streamlit as st
import pyexcel as p
import pandas as pd
from io import BytesIO
import re
from openpyxl import load_workbook


st.set_page_config(
    page_title="Monthly - App Interna",
    page_icon="💸",
    layout="wide"
)
primary_clr = st.get_option("theme.primaryColor")


st.markdown("""
<style>
button {
    background-color: #14E79D;
}
</style>
""", unsafe_allow_html=True)

header_logo_1, header_logo_2 = st.columns(2)
with header_logo_1:
    st.image(
                "./Logo. Monthly Oficial.png",
                width=250, # Manually Adjust the width of the image as per requirement
            )
with header_logo_2:
    st.markdown("<h2 style='text-align: right; color: #5666FF;'>✅ Procesador de Balanzas de Comprobación</h2>", unsafe_allow_html=True)

st.divider()


processor1, paddingA, processor2, paddingB, processor3 = st.columns([5,1,5,1,5])
with processor1:
    st.markdown("<h2 style='text-align: center; color: #14E79D; font-weight: bolder;'>PASO ☝️</h2>"+"<p style='text-align: center; color: #5666FF; font-weight: bold;'>Selecciona el ERP contable y sube el archivo</p>", unsafe_allow_html=True)
    st.markdown("")
    
    option = st.selectbox(
        "ERP:",
        ("Alpha ERP","Contpaqi","Contalink", "Microsip", "Aspel COI"))




    uploaded_file = st.file_uploader("Inserta el Excel (.xlsx, .csv) aquí:", type=['xlsx', 'xls'])

def process_data(df, option = option):
    tabs = list(df.keys())

#Microsip
    if option == "Microsip":
        outcome_df = pd.DataFrame(["","","",""], ["Cuenta", "Nombre", "Saldo Neto", "Sheet"]).T

        for tab in tabs:
            #Microsip Dataframe
            df = pd.read_excel(uploaded_file, sheet_name = tab)
            df = df.fillna(0)
            
            df.columns = df.loc[4]
            try:
                df = df.loc[5:].reset_index().drop(columns = {"index",0})
            except:
                df = df.loc[5:].reset_index()
                df = df.drop(columns = {"index"})
            
            type_of_Cuenta = []
            for i in list(df["Cuenta"]):
                type_of_Cuenta.append(type(i))
            
            df["Type"] = type_of_Cuenta
            df = df[df["Type"] == type("str")]
            df = df[:-1]
            df = df[df["Cuenta"] != "Fecha"]
            
            nivel_list = []
            for i in list(df["Cuenta"]):
                nivel_list.append(len(i.split(".")))
            
            df["Nivel"] = nivel_list
            class_of_cuenta = []
            for i in list(df["Cuenta"]):
                class_of_cuenta.append(i[0])
            
            df["Class"] = class_of_cuenta
            df["Class"] = df["Class"].astype("int")
            
            df.columns = ["Cuenta", "Nombre", "Saldo Inicial Deudor", "Saldo Inicial Acreedor", "Debe", "Haber", "Saldo Final Deudor", "Saldo Final Acreedor", "Tipo", "Nivel", "Clase"]
            df["Saldo Neto"] = df["Saldo Final Deudor"] - df["Saldo Final Acreedor"]
            
            #Microsip Validación
            activos = df[(df["Nivel"] == 1) & (df["Clase"] == 1)]["Saldo Neto"].sum()
            pasivos = df[(df["Nivel"] == 1) & (df["Clase"] == 2)]["Saldo Neto"].sum()
            capital= df[(df["Nivel"] == 1) & (df["Clase"] == 3)]["Saldo Neto"].sum()
            utilidad_acum = df[(df["Nivel"] == 1) & (df["Clase"] >= 4)]["Saldo Neto"].sum()
            
            result = round(activos+pasivos+capital+utilidad_acum,1)

            results_df = pd.DataFrame({"Cuentas": ["Activos", "Pasivos", "Capital", "Utilidad Acumulada", "Sumatoria Final"], 
                                    "Resultado":[activos, pasivos, capital, utilidad_acum, result]})
            
            if result == 0:
                st.caption(tab + " - Check ✅")
            else:
                st.caption(tab + " - No Check ❌")
            
            st.dataframe(results_df, width = 1000)
            st.divider()


                #Outcome Matrix
            
            balance_df = df[(df["Nivel"] == 2) & (df["Clase"] <= 3)][["Cuenta","Nombre","Saldo Neto"]]
            balance_df["Sheet"] = tab
        
            inc_statem_df = df[(df["Clase"] > 3)]
            inc_statem_df = inc_statem_df.drop(columns = ["Saldo Neto"])
            inc_statem_df["Saldo Neto"] = df["Haber"] - df["Debe"]
            general = []

            for i in inc_statem_df["Cuenta"]:
                general.append(i[:4])
            inc_statem_df["Cuenta General"] = general
            
            detalle_df = inc_statem_df.groupby(by = ["Cuenta General", "Nivel"])["Cuenta",].count().reset_index()
            
            nivel_deseado = []
            for c in list(set(detalle_df["Cuenta General"])):
                detalle_nivel_df = detalle_df[detalle_df["Cuenta General"] == c]
                
                if len(detalle_nivel_df[detalle_nivel_df["Cuenta"] >= 2]) > 0:
                    nivel_deseado.append([c, max(detalle_nivel_df[detalle_nivel_df["Cuenta"] >= 2]["Nivel"])])
                else:
                    nivel_deseado.append([c, 1])   
            nivel_deseado_df = pd.DataFrame(nivel_deseado, columns = ["Cuenta General", "Nivel Deseado"])
        
            detalle_deseado_df = detalle_df.merge(nivel_deseado_df, on = "Cuenta General", how = "left")[["Cuenta General", "Nivel Deseado"]].drop_duplicates()
            
            inc_statem_df = inc_statem_df.merge(detalle_deseado_df, on = "Cuenta General", how = "left")
            
            inc_statem_df = inc_statem_df[inc_statem_df["Nivel"] == inc_statem_df["Nivel Deseado"]]
            inc_statem_df = inc_statem_df[["Cuenta", "Nombre", "Saldo Neto"]]
            inc_statem_df["Sheet"] = tab
        
            outcome_df = pd.concat([outcome_df, balance_df, inc_statem_df])

        outcome_df = outcome_df.rename(columns = {"Cuenta":"Código", "Nombre":"Subcuenta"})
        tidy_df = outcome_df[["Código", "Subcuenta", "Sheet", "Saldo Neto"]]
        tidy_df = tidy_df[tidy_df["Código"].notnull()]
        tidy_df = tidy_df.dropna(how='all')
        tidy_df = tidy_df.iloc[1:]
        
        outcome_df = outcome_df.pivot(index=["Código", "Subcuenta"], columns=["Sheet"], values=["Saldo Neto"]).iloc[1:].reset_index().droplevel(0, axis = 1)
        outcome_df = outcome_df.fillna(0)

        outcome_columns = ["Código", "Subcuenta", "Espaciador"]

        for i in outcome_df.columns[3:]:
            outcome_columns.append(i)

        outcome_df.columns = outcome_columns
        outcome_df = outcome_df.drop(columns = ["Espaciador"])[["Código", "Subcuenta"] + tabs]
        outcome_df["Código"] = outcome_df["Código"].astype("str")

        return outcome_df, tidy_df, df

#Contpaqi
    if option == "Contpaqi":
        #Contpaqi Excel Sheets
        tabs = list(df.keys())
        tabs_dates = []

        outcome_df = pd.DataFrame(["","","",""], ["Cuenta", "Nombre", "Saldo Neto", "Sheet"]).T

        wb = load_workbook(uploaded_file)
        ws = wb.active  # Assuming we work with the first sheet
        
        # Check which cells in column A have bold formatting
        bold_cells = []
        for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, values_only=False):
            for cell in row:
                if cell.font and cell.font.bold:
                    bold_cells.append(cell.coordinate)
        
        
        # Initialize a dictionary to store bold cells for each sheet
        bold_cells_per_sheet = {}
        
        # Iterate over all sheets in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]  # Select the sheet
            bold_cells = []
        
            for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, values_only=False):
                for cell in row:
                    if cell.font and cell.font.bold:
                        bold_cells.append(cell.coordinate)
            
            bold_cells = [y.replace('A', '') for y in bold_cells]
            bold_cells = list(map(int, bold_cells))
            bold_cells = [x - 2 for x in bold_cells]
            bold_cells_per_sheet[sheet_name] = bold_cells
        


        for tab in tabs:
            df = pd.read_excel(uploaded_file, sheet_name = tab)
            df["Bold"] = df.index.isin(bold_cells_per_sheet[tab]).astype(int)
            tabs_date = df["CONTPAQ i"].iloc[0][-8:]
            tabs_dates.append(tabs_date)
            df = df.fillna(0)
            df = df.replace(' ',0)
            
            df.columns = ["Cuenta", "Nombre"]+list(df.loc[4][2:])
            df = df.iloc[6:-7].reset_index().drop(columns = ["index"])
            
            type_of_Cuenta = []
            for i in list(df["Cuenta"]):
                type_of_Cuenta.append(type(i))
                
            df["Type"] = type_of_Cuenta
            df = df[df["Type"] == type("str")]
            df = df[:-1]
            df = df[df["Cuenta"] != "Fecha"]
            
            nombre_list = []
            for n in df["Nombre"]:
                try:
                    nombre_list.append(n.strip())
                except:
                    nombre_list.append(n)
            
            df["Nombre"] = nombre_list

            
            nivel_list = []
            # Function to normalize ID by filling missing digits and sections
            def normalize_id(id_code):
                # Split the ID by hyphens
                sections = id_code.split('-')
                
                # Pad each section with zeros to ensure each section has exactly 4 digits
                padded_sections = [section.ljust(4, '0') for section in sections]
                
                # Ensure there are exactly 4 sections by appending "-0000" sections if needed
                while len(padded_sections) < 4:
                    padded_sections.append('0000')
                
                # Join the normalized sections with hyphens
                return '-'.join(padded_sections)
            
            # Function to determine the level based on the normalized ID
            def determine_nivel(id_code):
                # Normalize the ID first
                normalized_id = normalize_id(id_code)
                
                # Define patterns with corresponding levels
                patterns = {
                    r'^[1-9]000-0000-0000-0000$': 1,
                    r'^0[1-9]00-0000-0000-0000$': 2,
                    r'^[1-9]{2}00-0000-0000-0000$': 2,
                    r'^[1-9]0[1-9]0-0000-0000-0000$': 3,
                    r'^[1-9]0[1-9]{2}-0000-0000-0000$': 3,
                    r'^[1-9]00[1-9]-0000-0000-0000$': 3,
                    r'^0[1-9]{2}0-0000-0000-0000$': 3,
                    r'^0[1-9]0[1-9]-0000-0000-0000$': 3,
                    r'^0[1-9]{3}-0000-0000-0000$': 3,
                    r'^00[1-9]0-0000-0000-0000$': 3,
                    r'^000[1-9]-0000-0000-0000$': 3,
                    r'^00[1-9]{2}-0000-0000-0000$': 3,
                    r'^[1-9]{3}0-0000-0000-0000$': 3,
                    r'^[1-9]{2}0[1-9]-0000-0000-0000$': 3,
                    r'^[1-9]{4}-0000-0000-0000$': 3,
                    r'^\d{4}-[1-9]000-0000-0000$': 4,
                    r'^\d{4}-0[1-9]00-0000-0000$': 5,
                    r'^\d{4}-[1-9]{2}00-0000-0000$': 5,
                    r'^\d{4}-[1-9]0[1-9]0-0000-0000$': 6,
                    r'^\d{4}-[1-9]0[1-9]{2}-0000-0000$': 6,
                    r'^\d{4}-[1-9]00[1-9]-0000-0000$': 6,
                    r'^\d{4}-0[1-9]{2}0-0000-0000$': 6,
                    r'^\d{4}-0[1-9]0[1-9]-0000-0000$': 6,
                    r'^\d{4}-0[1-9]{3}-0000-0000$': 6,
                    r'^\d{4}-00[1-9]0-0000-0000$': 6,
                    r'^\d{4}-000[1-9]-0000-0000$': 6,
                    r'^\d{4}-00[1-9]{2}-0000-0000$': 6,
                    r'^\d{4}-[1-9]{3}0-0000-0000$': 6,
                    r'^\d{4}-[1-9]{2}0[1-9]-0000-0000$': 6,
                    r'^\d{4}-[1-9]{4}-0000-0000$': 6,
                    r'^\d{4}-\d{4}-[1-9]000-0000$': 7,
                    r'^\d{4}-\d{4}-0[1-9]00-0000$': 8,
                    r'^\d{4}-\d{4}-[1-9]{2}00-0000$': 8,
                    r'^\d{4}-\d{4}-[1-9]0[1-9]0-0000$': 9,
                    r'^\d{4}-\d{4}-[1-9]0[1-9]{2}-0000$': 9,
                    r'^\d{4}-\d{4}-[1-9]00[1-9]-0000$': 9,
                    r'^\d{4}-\d{4}-0[1-9]{2}0-0000$': 9,
                    r'^\d{4}-\d{4}-0[1-9]0[1-9]-0000$': 9,
                    r'^\d{4}-\d{4}-0[1-9]{3}-0000$': 9,
                    r'^\d{4}-\d{4}-00[1-9]0-0000$': 9,
                    r'^\d{4}-\d{4}-000[1-9]-0000$': 9,
                    r'^\d{4}-\d{4}-00[1-9]{2}-0000$': 9,
                    r'^\d{4}-\d{4}-[1-9]{3}0-0000$': 9,
                    r'^\d{4}-\d{4}-[1-9]{2}0[1-9]-0000$': 9,
                    r'^\d{4}-\d{4}-[1-9]{4}-0000$': 9,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]000$': 10,
                    r'^\d{4}-\d{4}-\d{4}-0[1-9]00$': 11,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]{2}00$': 11,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]0[1-9]0$': 12,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]0[1-9]{2}$': 12,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]00[1-9]$': 12,
                    r'^\d{4}-\d{4}-\d{4}-0[1-9]{2}0$': 12,
                    r'^\d{4}-\d{4}-\d{4}-0[1-9]0[1-9]$': 12,
                    r'^\d{4}-\d{4}-\d{4}-0[1-9]{3}$': 12,
                    r'^\d{4}-\d{4}-\d{4}-00[1-9]0$': 12,
                    r'^\d{4}-\d{4}-\d{4}-000[1-9]$': 12,
                    r'^\d{4}-\d{4}-\d{4}-00[1-9]{2}$': 12,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]{3}0$': 12,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]{2}0[1-9]$': 12,
                    r'^\d{4}-\d{4}-\d{4}-[1-9]{4}$': 12,
                }
                
                # Check the patterns and return the level if matched
                for pattern, level in patterns.items():
                    if re.match(pattern, normalized_id):
                        return level
                return None  # Return None if no pattern matches


            # Print results
            nivel_list = []
            cuenta_corr = []
            
            for i in list(df["Cuenta"]):
                cuenta_corr.append(normalize_id(i))
                nivel_list.append(determine_nivel(i))

        
            df['Cuenta'] = cuenta_corr
            df["Nivel"] = nivel_list
            
            class_of_cuenta = []
            for i in list(df["Cuenta"]):
                class_of_cuenta.append(i[0])
            
            df["Class"] = class_of_cuenta
            df["Class"] = df["Class"].astype("int")
            
            df.columns = ["Cuenta", "Nombre", "Saldo Inicial Deudor", "Saldo Inicial Acreedor", "Cargos", "Abonos", "Saldo Final Deudor", "Saldo Final Acreedor", "Bold", "Tipo", "Nivel", "Clase"]
            df["Saldo Final Deudor"] = df["Saldo Final Deudor"].astype("float")
            df["Saldo Final Acreedor"] = df["Saldo Final Acreedor"].astype("float")
            
            df["Saldo Neto"] = df["Saldo Final Deudor"] - df["Saldo Final Acreedor"]
            
            activos = df[(df["Bold"] == 1) & (df["Clase"] == 1)]["Saldo Neto"].sum()
            pasivos = df[(df["Bold"] == 1) & (df["Clase"] == 2)]["Saldo Neto"].sum()
            capital= df[(df["Bold"] == 1) & (df["Clase"] == 3)]["Saldo Neto"].sum()


            utilidad_acum = df[(df["Nivel"] == min(df[df["Clase"] >= 4]["Nivel"])) & (df["Clase"] >= 4)]["Saldo Neto"].sum()
            
            result = round(activos+pasivos+capital+utilidad_acum,1)
            
            results_df = pd.DataFrame({"Cuentas": ["Activos", "Pasivos", "Capital", "Utilidad Acumulada", "Sumatoria Final"], 
                                    "Resultado":[activos, pasivos, capital, utilidad_acum, result]})
            
            if result == 0:
                st.caption(tabs_date + " - Check ✅")
            else:
                st.caption(tabs_date + " - No Check ❌")
            
            st.dataframe(results_df, width = 1000)
            st.divider()
            
            
            #Outcome Matrix
            
            df_unique = df.drop_duplicates(subset="Código", keep="first").copy()
            df_unique = df_unique.rename(columns = {"Cuenta":"Código", "Nombre":"Cuenta"])
            cuentas_dict = df_unique.set_index("Código").to_dict(orient="index")
            
            # === 3. Función para obtener el padre jerárquico correcto ===
            def obtener_padre(cuenta, todas_cuentas):
                partes = cuenta.split('-')
                for i in reversed(range(1, len(partes))):
                    posible_padre = '-'.join(partes[:i] + ['0000'] * (4 - i))
                    if posible_padre in todas_cuentas and posible_padre != cuenta:
                        return posible_padre
                return None
            
            # === 4. Calcular columnas jerárquicas ===
            cuentas_existentes = set(df_unique["Código"])
            df["Padre"] = df["Código"].apply(lambda x: obtener_padre(x, cuentas_existentes))
            df["Cuenta del Padre"] = df["Padre"].map(lambda x: cuentas_dict[x]["Cuenta"] if x in cuentas_dict else None)
            df["Nivel"] = df["Código"].apply(lambda x: sum(1 for seg in x.split("-") if seg != '0000'))
            # === 5. Calcular suma de subcuentas para cada padre ===
            suma_subcuentas = (
                df.groupby("Padre")["Saldo Neto"]
                .sum()
                .reset_index()
                .rename(columns={"Padre": "Código Padre", "Saldo Neto": "Suma Hijos"})
            )
            
            # === 6. Unir datos originales con sumas de hijos ===
            df_validado = df.merge(suma_subcuentas, how="left", left_on="Código", right_on="Código Padre")
            
            # === 7. Verificar si cuadran subcuentas con el saldo del padre ===
            df_validado["Subcuentas cuadran"] = df_validado.apply(
                lambda row: (
                    np.isclose(row["Saldo Neto"], row["Suma Hijos"], atol=1)
                    if not pd.isna(row["Suma Hijos"])
                    else False
                ),
                axis=1,
            )
            
            # === 8. Verificar si la cuenta tiene hijos directos ===
            def tiene_hijos(cuenta):
                return any(obtener_padre(child, cuentas_existentes) == cuenta for child in cuentas_existentes)
            
            df_validado["Tiene Hijos"] = df_validado["Código"].apply(tiene_hijos)
            
            # === 9. Clasificar el estado de la cuenta ===
            def clasificar_estado(row):
                if row["Subcuentas cuadran"]:
                    return "Cuadran"
                elif not row["Tiene Hijos"]:
                    return "No subcuentas"
                else:
                    return "No cuadran"
            
            df_validado["Estado"] = df_validado.apply(clasificar_estado, axis=1)
            
            df_completo = df_validado[[
                "Código", "Cuenta", "Padre", "Nombre del Padre", "Nivel",
                "Saldo Neto", "Suma Hijos", "Subcuentas cuadran", "Tiene Hijos", "Estado"
            ]]
            
            df_completo_ajuste = pd.DataFrame()
            
            for i in list(df_completo["Código"].str[0].unique()):
                df_seccion = df_completo[df_completo["Código"].str.startswith(i)]
                sección = df_seccion["Cuenta"].iloc[0]
                cuenta = df_seccion["Código"].iloc[0]
                valores = df_seccion[["Padre", "Cuenta del Padre"]].iloc[0:2].values.flatten()
                todos_none = all(v is None for v in valores)
                nulos = df_seccion[df_seccion["Padre"].isnull()].iloc[1:]
            
                if todos_none:  # All None
                    df_seccion["Cuenta del Padre"].iloc[1:] = df_seccion["Cuenta del Padre"].iloc[1:].fillna(sección)
                    df_seccion["Padre"].iloc[1:] = df_seccion["Padre"].iloc[1:].fillna(cuenta)
            
                    primer_cuenta = list(df_seccion.index)[0]
                    #df_seccion.loc[primer_cuenta + 1:, "Nivel"] += 1
                    df_seccion.loc[primer_cuenta, "Nivel"] -= 1
                    df_seccion["Nivel"] += 1
                    
                else: #Not None
                    padres = df_seccion[df_seccion["Cuenta del Padre"] == sección]
                    padres_list = list(padres["Cuenta"])
                    cuenta_list = list(padres["Código"])
                    nivel_list = list(padres["Nivel"])
                    index_list = list(padres.index)
                    for x in range(len(index_list)):
                        padre = padres_list[x]
                        cuenta = cuenta_list[x]
                        nivel = nivel_list[x]
                        try:
                            df_seccion.loc[index_list[x]:index_list[x+1], "Cuenta del Padre"] = df_seccion.loc[index_list[x]:index_list[x+1], "Cuenta del Padre"].fillna(padre)
                            df_seccion.loc[index_list[x]:index_list[x+1], "Padre"] = df_seccion.loc[index_list[x]:index_list[x+1], "Padre"].fillna(cuenta)
                        except:
                            df_seccion.loc[index_list[x]:list(df_seccion.index)[-1], "Cuenta del Padre"] = df_seccion.loc[index_list[x]:list(df_seccion.index)[-1], "Cuenta del Padre"].fillna(padre)
                            df_seccion.loc[index_list[x]:list(df_seccion.index)[-1], "Padre"] = df_seccion.loc[index_list[x]:list(df_seccion.index)[-1], "Padre"].fillna(cuenta)
                        
                df_completo_ajuste = pd.concat([df_completo_ajuste, df_seccion])
            
            df_completo_ajuste = df_completo_ajuste.drop(columns = "Suma Hijos")
            
            nivel_list = []
            
            counter = 0
            for _, row in df_completo_ajuste.iterrows():
                if row["Padre"] is None or row["Padre"] == "":
                    df_completo_ajuste["Nivel"].iloc[counter] = 1
                else:
                    df_completo_ajuste["Nivel"].iloc[counter] = df_completo_ajuste[df_completo_ajuste["Código"] == row["Padre"]]["Nivel"].values[0] + 1
                counter += 1
                                        
            
            # === 5. Calcular suma de subcuentas para cada padre ===
            suma_subcuentas = (
                df_completo_ajuste.groupby("Padre")["Saldo Neto"]
                .sum()
                .reset_index()
                .rename(columns={"Padre": "Código Padre", "Saldo Neto": "Suma Hijos"})
            )
            
            # === 6. Unir datos originales con sumas de hijos ===
            df_final = df_completo_ajuste.merge(suma_subcuentas, how="left", left_on="Código", right_on="Código Padre")
            
            # === 7. Verificar si cuadran subcuentas con el saldo del padre ===
            df_final["Subcuentas cuadran"] = df_final.apply(
                lambda row: (
                    np.isclose(row["Saldo Neto"], row["Suma Hijos"], atol=1)
                    if not pd.isna(row["Suma Hijos"])
                    else False
                ),
                axis=1,
            )
            
            # === 8. Verificar si la cuenta tiene hijos directos ===
            def tiene_hijos(cuenta):
                if len(df_final[df_final["Padre"] == cuenta]) != 0:
                    return True
                else:
                    return False
            
            df_final["Tiene Hijos"] = df_final["Código"].apply(tiene_hijos)
            
            # === 8. Clasificar el estado de la cuenta ===
            def clasificar_estado(row):
                if row["Subcuentas cuadran"]:
                    return "Cuadran"
                elif not row["Tiene Hijos"]:
                    return "No subcuentas"
                else:
                    return "No cuadran"
            
            df_final["Estado"] = df_final.apply(clasificar_estado, axis=1)
            
            # === 9. Tabla final con todo lo necesario ===
            df_final_nivel = df_final[[
                "Código", "Cuenta", "Padre", "Nombre del Padre", "Bold", "Cargos", "Abonos", "Saldo Neto", "Nivel", "Clase"]]
            
            balance_df = df[(df["Bold"] == 1) & (df["Clase"] <= 3)][["Cuenta","Nombre","Saldo Neto"]]
            balance_df["Sheet"] = tabs_date

            inc_statem_df = df[(df["Clase"] > 3)]
            inc_statem_df = inc_statem_df.drop(columns = ["Saldo Neto"])
            inc_statem_df["Saldo Neto"] = inc_statem_df["Cargos"] - inc_statem_df["Abonos"]
            inc_statem_df = inc_statem_df[["Cuenta", "Nombre", "Saldo Neto"]]
            inc_statem_df["Sheet"] = tabs_date

            outcome_df = pd.concat([outcome_df, balance_df, inc_statem_df])

        outcome_df = outcome_df.rename(columns = {"Cuenta":"Subcuenta"})
        tidy_df = outcome_df[["Código", "Subcuenta", "Sheet", "Saldo Neto"]]
        tidy_df = tidy_df[tidy_df["Código"].notnull()]
        tidy_df = tidy_df.dropna(how='all')
        tidy_df = tidy_df.iloc[1:]
        
        outcome_df = outcome_df.pivot(index=["Código", "Subcuenta", "Padre", "Nombre del Padre", columns=["Sheet"], values=["Saldo Neto"]).iloc[1:].reset_index().droplevel(0, axis = 1)
        outcome_df = outcome_df.fillna(0)

        outcome_columns = ["Código", "Subcuenta", "Padre", "Nombre del Padre"]

        for i in outcome_df.columns[2:]:
            outcome_columns.append(i)

        outcome_df.columns = outcome_columns
        outcome_df = outcome_df[["Código", "Subcuenta", "Padre", "Nombre del Padre"] + tabs_dates]

        outcome_df["Código"] = outcome_df["Código"].astype("str")
        
        return outcome_df, tidy_df, df

#Contalink
    if option == "Contalink":
        #Contalink Excel Sheets
        tabs = list(df.keys())
        tabs_dates = []

        outcome_df = pd.DataFrame(["","","",""], ["Cuenta", "Nombre", "Saldo Neto", "Sheet"]).T


        for tab in tabs:
            df = pd.read_excel(uploaded_file, sheet_name = tab)
            tabs_date = df.iloc[1,2][-7:]
            tabs_dates.append(tabs_date)
            
            df.columns = df.iloc[2]
            df = df.dropna(subset = df.columns[1])
            df = df[["No CUENTA","CUENTA", "SALDO INICIAL", "DEBE", "HABER", "SALDO FINAL"]]
            
            class_of_cuenta = []
            for i in list(df["No CUENTA"]):
                class_of_cuenta.append(i[0])
            
            df["Class"] = class_of_cuenta
            df["Class"] = df["Class"].astype("int")
            
            codigo_verif = []
            for i in df["No CUENTA"]:
                codigo_verif.append(len(i.split("-")))
            
            df["Código Verif"] = codigo_verif
            df = df[df["Código Verif"] == 3].drop(columns = "Código Verif")
            
            
            nivel_list = []
                
            for i in list(df["No CUENTA"]):
                if (i.split("-")[2] == '000'):
                    if (i.split("-")[1] == '000'):
                        nivel_list.append(1)
                    else:
                        nivel_list.append(2)
                else:
                    nivel_list.append(3)
            
            df["Nivel"] = nivel_list
                
            df.columns = ["Código", "Cuenta", "Saldo Inicial", "Debe", "Haber", "Saldo Final", "Clase", "Nivel"]
            #df["Saldo Neto"] = df["Saldo Neto"].str.replace(',', '')
            df["Saldo Final"] = df["Saldo Final"].astype(float)
        
            #Depreciacion Corregida
            #df[df["Cuenta"] == 'DEPRECIACION ACUMULADA'] = -1*df[df["Cuenta"] == 'DEPRECIACION ACUMULADA'][["Saldo Inicial", "Debe", "Haber", "Saldo Neto"]]
            df.loc[df["Cuenta"] == 'DEPRECIACION ACUMULADA', ["Saldo Inicial", "Debe", "Haber", "Saldo Final"]] = df.loc[df["Cuenta"] == 'DEPRECIACION ACUMULADA', ["Saldo Inicial", "Debe", "Haber", "Saldo Final"]] * -1
        
            saldo_neto = []
            for i in range(len(df)):
                if df["Clase"].iloc[i] <= 3:
                    saldo_neto.append(df["Saldo Final"].iloc[i])
                else:
                    saldo_neto.append(df["Haber"].iloc[i]-df["Debe"].iloc[i])
        
            df["Saldo Neto"] = saldo_neto
        
            activos = df[(df["Nivel"] == 1) & (df["Clase"] == 1)]["Saldo Neto"].sum()
            pasivos = df[(df["Nivel"] == 1) & (df["Clase"] == 2)]["Saldo Neto"].sum()
            capital = df[(df["Nivel"] == 1) & (df["Clase"] == 3)]["Saldo Neto"].sum()
            utilidad_acum = df[(df["Nivel"] == 1) & (df["Clase"] >= 4)]["Saldo Neto"].sum()
            
            result = round(activos-pasivos-capital-utilidad_acum,1)
                
            results_df = pd.DataFrame({"Cuentas": ["Activos", "Pasivos", "Capital", "Utilidad Acumulada", "Sumatoria Final"], 
                                       "Resultado":[activos, pasivos, capital, utilidad_acum, result]})
            
            if result == 0:
                st.caption(tabs_date + " - Check ✅")
            else:
                st.caption(tabs_date + " - No Check ❌")
            
            st.dataframe(results_df, width = 1000)
            st.divider()


            #Outcome Matrix
            balance_df = df[(df["Nivel"] == 1) & (df["Clase"] <= 3)][["Código","Cuenta","Saldo Neto"]]
            balance_df["Sheet"] = tabs_date
        
            inc_statem_df = df[(df["Clase"] > 3)]
            general = []
            
            for i in inc_statem_df["Código"]:
                general.append(i[:4])
            inc_statem_df["Cuenta General"] = general
            
            detalle_df = inc_statem_df.groupby(by = ["Cuenta General", "Nivel"]).agg({"Código":"count", "Saldo Neto":"sum"}).reset_index()
            
            nivel_deseado = []
            for c in list(set(detalle_df["Cuenta General"])):
                detalle_nivel_df = detalle_df[detalle_df["Cuenta General"] == c]    
                if len(detalle_nivel_df[detalle_nivel_df["Código"] >= 2]) > 0:
                    nivel_deseado.append([c, max(detalle_nivel_df[detalle_nivel_df["Código"] >= 2]["Nivel"])])
                else:
                    nivel_deseado.append([c, 1])   
            nivel_deseado_df = pd.DataFrame(nivel_deseado, columns = ["Cuenta General", "Nivel Deseado"])
            
            detalle_deseado_df = detalle_df.merge(nivel_deseado_df, on = "Cuenta General", how = "left")[["Cuenta General", "Nivel Deseado"]].drop_duplicates()
            
            inc_statem_df = inc_statem_df.merge(detalle_deseado_df, on = "Cuenta General", how = "left")
            
            inc_statem_df = inc_statem_df[inc_statem_df["Nivel"] == inc_statem_df["Nivel Deseado"]]
            inc_statem_df = inc_statem_df[["Código", "Cuenta", "Saldo Neto"]]
            inc_statem_df["Sheet"] = tabs_date
            
            outcome_df = pd.concat([outcome_df, balance_df, inc_statem_df])

        outcome_df = outcome_df.rename(columns = {"Cuenta":"Subcuenta"})
        tidy_df = outcome_df[["Código", "Subcuenta", "Sheet", "Saldo Neto"]]
        tidy_df = outcome_df
        outcome_df = outcome_df.pivot(index=["Código", "Subcuenta"], columns=["Sheet"], values=["Saldo Neto"]).iloc[1:].reset_index().droplevel(0, axis = 1)
        outcome_df = outcome_df.fillna(0)
        tidy_df = tidy_df.iloc[1:]
        tidy_df = tidy_df.drop(columns = "Nombre")        
        
        outcome_columns = ["Código", "Subcuenta"]
        
        for i in outcome_df.columns[2:]:
            outcome_columns.append(i)
        
        outcome_df.columns = outcome_columns
        outcome_df = outcome_df[["Código", "Subcuenta"] + tabs_dates]
        
        outcome_df["Código"] = outcome_df["Código"].astype("str")
        
        return outcome_df, tidy_df, df

    
#Aspel COI
    if option == "Aspel COI":
        #Contpaqi Excel Sheets
        tabs = list(df.keys())
        tabs_dates = []

        outcome_df = pd.DataFrame(["","","",""], ["Cuenta", "Nombre", "Saldo Neto", "Sheet"]).T


        for tab in tabs:
            df = pd.read_excel(uploaded_file, sheet_name = tab)
            df.columns = df.iloc[11]
            df = df.dropna(subset = df.columns[1])[2:]
            tabs_dates.append(tab)
            
            column_list = list(df.columns)
            column_list = [x for x in column_list if str(x) != 'nan']
            df = df[column_list]
            
            cuenta = []
            cuenta_len = []
            for i in list(df['No. de cuenta       Descripción']):
                cuenta.append([x for x in i.split(" ") if str(x) != ""])
                cuenta_len.append(len([x for x in i.split(" ") if str(x) != ""]))
            
            df["Cuenta"] = cuenta
            df["Validación"] = cuenta_len
            df = df[df["Validación"] > 1].drop(columns = [df.columns[0], "Validación"])
            
            codigo = []
            nombre = []
            for i in list(df["Cuenta"]):
                if len(i[0]) == 13:
                    codigo.append(i[0])
                else:
                    codigo.append(0)
                nombre.append(" ".join(str(n) for n in i[1:]))
            
            df["Cuenta"] = codigo
            df["Nombre"] = nombre
            df = df[df["Cuenta"] != 0]
            
            nivel_list = []
            
            for i in list(df["Cuenta"]):
                if (i.split("-")[1] == '0000'):
                    nivel_list.append(1)
                elif (i.split("-")[2][-2:] == '000') and (i.split("-")[1] != '0000'):
                    nivel_list.append(2)
                else:
                    nivel_list.append(3)
                        
            
            df["Nivel"] = nivel_list
            
            class_of_cuenta = []
            for i in list(df["Cuenta"]):
                class_of_cuenta.append(i[0])
            
            df["Class"] = class_of_cuenta
            df["Class"] = df["Class"].astype("int")
            
            df.columns = ["Saldo Inicial", "Debe", "Haber", "Saldo Neto", "Cuenta", "Nombre", "Nivel", "Clase"]
            #df["Saldo Final Deudor"] = df["Saldo Final Deudor"].astype("float")
            #df["Saldo Final Acreedor"] = df["Saldo Final Acreedor"].astype("float")
            df["Saldo Neto"] = df["Saldo Neto"].str.replace(',', '')
            df["Saldo Neto"] = df["Saldo Neto"].astype(float)
            
            
            activos = df[(df["Nivel"] == 1) & (df["Clase"] == 1)]["Saldo Neto"].sum()
            pasivos = df[(df["Nivel"] == 1) & (df["Clase"] == 2)]["Saldo Neto"].sum()
            capital= df[(df["Nivel"] == 1) & (df["Clase"] == 3)]["Saldo Neto"].sum()
            utilidad_acum = df[(df["Nivel"] == 1) & (df["Clase"] >= 4)]["Saldo Neto"].sum()
            
            result = round(activos+pasivos+capital+utilidad_acum,1)
                
            results_df = pd.DataFrame({"Cuentas": ["Activos", "Pasivos", "Capital", "Utilidad Acumulada", "Sumatoria Final"], 
                                       "Resultado":[activos, pasivos, capital, utilidad_acum, result]})
        
            if result == 0:
                st.caption(tab + " - Check ✅")
            else:
                st.caption(tab + " - No Check ❌")
            
            st.dataframe(results_df, width = 1000)
            st.divider()

             #Outcome Matrix
            balance_df = df[(df["Nivel"] == 1) & (df["Clase"] <= 3)][["Cuenta","Nombre","Saldo Neto"]]
            balance_df["Sheet"] = tab
        
            inc_statem_df = df[(df["Clase"] > 3)]
            general = []
            
            for i in inc_statem_df["Cuenta"]:
                general.append(i[:4])
            inc_statem_df["Cuenta General"] = general
            
            detalle_df = inc_statem_df.groupby(by = ["Cuenta General", "Nivel"]).agg({"Cuenta":"count", "Saldo Neto":"sum"}).reset_index()
            
            nivel_deseado = []
            for c in list(set(detalle_df["Cuenta General"])):
                detalle_nivel_df = detalle_df[detalle_df["Cuenta General"] == c]    
                if len(detalle_nivel_df[detalle_nivel_df["Cuenta"] >= 2]) > 0:
                    nivel_deseado.append([c, max(detalle_nivel_df[detalle_nivel_df["Cuenta"] >= 2]["Nivel"])])
                else:
                    nivel_deseado.append([c, 1])   
            nivel_deseado_df = pd.DataFrame(nivel_deseado, columns = ["Cuenta General", "Nivel Deseado"])
            
            detalle_deseado_df = detalle_df.merge(nivel_deseado_df, on = "Cuenta General", how = "left")[["Cuenta General", "Nivel Deseado"]].drop_duplicates()
            
            inc_statem_df = inc_statem_df.merge(detalle_deseado_df, on = "Cuenta General", how = "left")
            
            inc_statem_df = inc_statem_df[inc_statem_df["Nivel"] == inc_statem_df["Nivel Deseado"]]
            inc_statem_df = inc_statem_df[["Cuenta", "Nombre", "Saldo Neto"]]
            inc_statem_df["Sheet"] = tab
            
            outcome_df = pd.concat([outcome_df, balance_df, inc_statem_df])

        outcome_df = outcome_df.rename(columns = {"Cuenta":"Código" ,"Nombre":"Subcuenta"})
        tidy_df = outcome_df[["Código", "Subcuenta", "Sheet", "Saldo Neto"]]
        tidy_df = tidy_df[tidy_df["Subcuenta"].notnull()]
        tidy_df = tidy_df.dropna(how='all')
        tidy_df = tidy_df.iloc[1:]
        
        outcome_df = outcome_df.pivot(index=["Código", "Subcuenta"], columns=["Sheet"], values=["Saldo Neto"]).iloc[1:].reset_index().droplevel(0, axis = 1)
        outcome_df = outcome_df.fillna(0)
        
        outcome_columns = ["Código", "Subcuenta"]
        
        for i in outcome_df.columns[2:]:
            outcome_columns.append(i)
        
        outcome_df.columns = outcome_columns
        outcome_df = outcome_df[["Código", "Subcuenta"] + tabs_dates]
        
        outcome_df["Código"] = outcome_df["Código"].astype("str")
        
        return outcome_df, tidy_df, df

    

    if option == "Alpha ERP":
        ###### Alpha Excel Sheets
        tabs = list(df.keys())
        tabs_dates = []

        outcome_df = pd.DataFrame(["","","",""], ["Cuenta", "Nombre", "Saldo Neto", "Sheet"]).T
        
        #Alpha Excel Sheets
        for tab in tabs:
            df = pd.read_excel(uploaded_file, sheet_name = tab)
            tabs_date = df.iloc[1, 5][-8:]
            tabs_dates.append(tabs_date)
            
            df = df.fillna(0)[6:-4]
            
            zero_cols = df.columns[(df == 0).all()]
            df.drop(labels=zero_cols, axis=1, inplace=True)
            df.columns = ["Código", "Cuenta", "Saldo Inicial Deudor", "Saldo Inicial Acreedor", "Cargos", "Abonos", "Saldo Final Deudor", "Saldo Final Acreedor"]
            df = df[df["Código"] != 0]
            
            nivel_list = []
            def count_leading_spaces(s):
                count = 0
                for char in s:
                    if char == ' ':
                        count += 1
                    else:
                        break
                return count
            
            for i in df["Cuenta"]:
                nivel_list.append(int(1+count_leading_spaces(i)/2))
            
            df["Nivel"] = nivel_list
            
            class_of_cuenta = []
            for i in list(df["Código"]):
                class_of_cuenta.append(i[0])
            
            df["Clase"] = class_of_cuenta
            df["Clase"] = df["Clase"].astype("int")
            
            df["Saldo Neto"] = df["Saldo Final Deudor"] - df["Saldo Final Acreedor"]

            validacion = []
            for i in df["Código"]:
                validacion.append(i[-1].isdigit())
            
            df["Valid"] = validacion
            df = df[(df["Valid"] == True) | ((df["Valid"] == False) & (df["Nivel"] == 2))]
            
            activos = df[(df["Nivel"] == 2) & (df["Clase"] == 1)]["Saldo Neto"].sum()
            pasivos = df[(df["Nivel"] == 2) & (df["Clase"] == 2)]["Saldo Neto"].sum()
            capital= df[(df["Nivel"] == 2) & (df["Clase"] == 3)]["Saldo Neto"].sum()
            utilidad_acum = df[(df["Nivel"] == 2) & (df["Clase"] >= 4)]["Saldo Neto"].sum()
            
            result = round(activos+pasivos+capital+utilidad_acum,1)
                
            results_df = pd.DataFrame({"Cuentas": ["Activos", "Pasivos", "Capital", "Utilidad Acumulada", "Sumatoria Final"], 
                                       "Resultado":[activos, pasivos, capital, utilidad_acum, result]})
            
            if result == 0:
                st.caption(tabs_date + " - Check ✅")
            else:
                st.caption(tabs_date + " - No Check ❌")
            
            st.dataframe(results_df, width = 1000)
            st.divider()


            #Outcome Matrix
            balance_df = df[(df["Nivel"] == 3) & (df["Clase"] <= 3)][["Código","Cuenta","Saldo Neto"]]
            balance_df["Sheet"] = tabs_date
        
            inc_statem_df = df[(df["Clase"] > 3)]
            general = []
            
            for i in inc_statem_df["Código"]:
                general.append(i[:4])
            inc_statem_df["Cuenta General"] = general
            
            detalle_df = inc_statem_df.groupby(by = ["Cuenta General", "Nivel"]).agg({"Código":"count", "Saldo Neto":"sum"}).reset_index()
            
            nivel_deseado = []
            for c in list(set(detalle_df["Cuenta General"])):
                detalle_nivel_df = detalle_df[detalle_df["Cuenta General"] == c]    
                if len(detalle_nivel_df[detalle_nivel_df["Código"] >= 2]) > 0:
                    nivel_deseado.append([c, max(detalle_nivel_df[detalle_nivel_df["Código"] >= 2]["Nivel"])])
                else:
                    nivel_deseado.append([c, 1])   
            nivel_deseado_df = pd.DataFrame(nivel_deseado, columns = ["Cuenta General", "Nivel Deseado"])
            
            detalle_deseado_df = detalle_df.merge(nivel_deseado_df, on = "Cuenta General", how = "left")[["Cuenta General", "Nivel Deseado"]].drop_duplicates()
        
            inc_statem_df = inc_statem_df.merge(detalle_deseado_df, on = "Cuenta General", how = "left")
            
            inc_statem_df = inc_statem_df[inc_statem_df["Nivel"] == inc_statem_df["Nivel Deseado"]]
            inc_statem_df = inc_statem_df[["Código","Cuenta","Saldo Neto"]]
            inc_statem_df["Sheet"] = tabs_date
            
            outcome_df = pd.concat([outcome_df, balance_df, inc_statem_df])

        outcome_df = outcome_df.rename(columns = {"Cuenta":"Subcuenta"})
        tidy_df = outcome_df[["Código", "Subcuenta", "Sheet", "Saldo Neto"]]
        tidy_df = tidy_df[tidy_df["Código"].notnull()]
        tidy_df = tidy_df.dropna(how='all')
        tidy_df = tidy_df.rename(columns = {"Cuenta":"Subcuenta"})

        
        outcome_df = outcome_df.pivot(index=["Código","Subcuenta"], columns=["Sheet"], values=["Saldo Neto"]).iloc[1:].reset_index().droplevel(0, axis = 1)
        outcome_df = outcome_df.fillna(0)
        
        outcome_columns = ["Código","Subcuenta"]
        
        for i in outcome_df.columns[2:]:
            outcome_columns.append(i)
        
        outcome_df.columns = outcome_columns
        outcome_df = outcome_df[["Código","Subcuenta"] + tabs_dates]
        
        outcome_df["Código"] = outcome_df["Código"].astype("str")
        
        return outcome_df, tidy_df, df


if uploaded_file is not None:
    
    if uploaded_file.name[-3:] == "xls":
        try:
            p.save_book_as(file_name = uploaded_file.name, dest_file_name = uploaded_file.name[:-4] + '.xlsx')
        except:
            pass
    
    df = pd.read_excel(uploaded_file, sheet_name = None)
    
    if st.button('Correr 🚶‍♂️'):
        # Process the DataFrame
        with processor2:
            st.markdown("<h2 style='text-align: center; color: #14E79D; font-weight: bolder;'>PASO ✌️</h2>"+"<p style='text-align: center; color: #5666FF; font-weight: bold;'>Valida que la información cuadre</p>", unsafe_allow_html=True)
            st.markdown("")
            processed_df = process_data(df)
            
            # Convert DataFrame to Excel
            matrix = BytesIO()
            processed_df[0].to_excel(matrix, index=False, engine = 'openpyxl')  # Using 'openpyxl' for .xlsx files
            matrix.seek(0)

            tidy = BytesIO()
            processed_df[1].to_excel(tidy, index=False, engine = 'openpyxl')  # Using 'openpyxl' for .xlsx files
            tidy.seek(0)

            details = BytesIO()
            processed_df[2].to_excel(details, index=False, engine = 'openpyxl')  # Using 'openpyxl' for .xlsx files
            details.seek(0)

            st.download_button(label='Descarga Tabla de Detalle 📃',
                            data=details,
                            file_name='details.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            

        
            

        with processor3:
            st.markdown("<h2 style='text-align: center; color: #14E79D; font-weight: bolder;'>PASO 👌</h2>"+"<p style='text-align: center; color: #5666FF; font-weight: bold;'>Descarga el archivo procesado</p>", unsafe_allow_html=True)
            st.markdown("")
            # Optionally display the processed DataFrame (comment out if not needed)
            st.dataframe(processed_df[0])

            # Create a link for downloading
            st.download_button(label='Descarga Matriz 📃',
                            data=matrix,
                            file_name='matrix_output.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            st.markdown("")
            st.divider()
            st.markdown("")
            
            st.dataframe(processed_df[1])

            # Create a link for downloading
            st.download_button(label='Descarga Tabular 📃',
                            data=tidy,
                            file_name='tidy_output.xlsx',
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
