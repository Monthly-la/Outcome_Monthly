import streamlit as st
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Border, Fill, Alignment, Protection

st.set_page_config(
    page_title="Monthly - App Interna",
    page_icon="💸",
    layout="wide"
)

st.markdown("""
<style>
button {
    background-color: #14E79D;
}
</style>
""", unsafe_allow_html=True)

header_logo_1, header_logo_2 = st.columns(2)
with header_logo_1:
    st.image("./Logo. Monthly Oficial.png", width=250)
with header_logo_2:
    st.markdown("<h2 style='text-align: right; color: #5666FF;'>📑 Recopilador de Documentos</h2>", unsafe_allow_html=True)

st.divider()

def merge_excel_files(excel_files):
    output = BytesIO()
    
    # Create a new workbook to store merged sheets
    merged_wb = openpyxl.Workbook()
    merged_wb.remove(merged_wb.active)  # Remove default sheet
    
    for uploaded_file in excel_files:
        # Load the uploaded workbook
        wb = openpyxl.load_workbook(uploaded_file)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Create a new sheet in the merged workbook
            new_sheet = merged_wb.create_sheet(title=sheet_name[:31])  # Sheet names are limited to 31 chars
            
            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)

                    # Copy styles safely
                    try:
                        new_cell.font = cell.font.copy() if cell.font else None
                        new_cell.border = cell.border.copy() if cell.border else None
                        new_cell.fill = cell.fill.copy() if cell.fill else None
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy() if cell.protection else None
                        new_cell.alignment = cell.alignment.copy() if cell.alignment else None
                    except Exception as e:
                        print(f"Error copying styles in cell {cell.coordinate}: {e}")

    # Save the final workbook
    merged_wb.save(output)
    output.seek(0)  # Reset pointer for reading
    return output

excel_files = st.file_uploader("Inserta el Excel (.xlsx, .csv) aquí:", type=['xlsx'], accept_multiple_files=True)

if excel_files:
    if st.button("Merge Files"):
        merged_file = merge_excel_files(excel_files)
        
        st.success("Files have been merged successfully!")

        st.download_button(
            label="Download Merged Excel",
            data=merged_file.getvalue(),
            file_name="merged_file.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
